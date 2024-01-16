
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name = []

# program read information from people.csv file and put all data in name list.
with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row = line.rstrip().split(",")
        name.append(str(row[2]+" " + str(row[3])))


url = "https://emn178.github.io/online-tools/crc32.html"


decoded_workbook = Workbook()
decoded_sheet = decoded_workbook.active
decoded_sheet.title = "Decoded Information"
decoded_sheet.append(["Employee Name", "Encoded Name"])


salary_file = load_workbook("salary.xlsx")
salary_sheet = salary_file.active

driver.get(url)
time.sleep(2)
for employee in name:
    full_name = employee  


   
    input_element = driver.find_element(By.ID, 'input')
    input_element.clear()
    input_element.send_keys(full_name)

   
    encoded_result = driver.find_element(By.ID, 'output').get_attribute('value')

    
    decoded_sheet.append([full_name, encoded_result])


decoded_workbook.save("decoded_information.xlsx")


driver.quit()